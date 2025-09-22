#!/usr/bin/env python3
"""
Script to generate test Excel files for all 5 test scenarios
"""

import sys
import os
from pathlib import Path

# Add the current directory to the path to import modules
sys.path.insert(0, os.path.dirname(__file__))

try:
    import openpyxl
    from openpyxl import Workbook
    print("✅ Successfully imported required modules")
except ImportError as e:
    print(f"❌ Failed to import modules: {e}")
    print("Please install openpyxl: pip install openpyxl")
    sys.exit(1)

def create_stability_file(filename, scenario):
    """Create stability analysis Excel file for specific scenario"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise Exception("Failed to create worksheet")
    
    ws.title = "Stability Analysis"
    
    # Headers
    ws['A1'] = f'{scenario["Bridge Name"]} - Stability Analysis'
    ws['A3'] = 'Project Parameters'
    ws['A5'] = 'Parameter'
    ws['B5'] = 'Value'
    ws['C5'] = 'Unit'
    
    # Parameters based on scenario
    params = [
        ('Structure Height', scenario.get('Effective Span', 10) * 0.6, 'm'),
        ('Structure Width', scenario.get('Bridge Width', 7.5), 'm'),
        ('Concrete Density', 24.0, 'kN/m³'),
        ('Soil Unit Weight', 18.0, 'kN/m³'),
        ('Angle of Friction', 30.0, 'degrees'),
        ('Bearing Capacity', 450.0, 'kN/m²')
    ]
    
    for i, (param, value, unit) in enumerate(params, start=6):
        ws[f'A{i}'] = param
        ws[f'B{i}'] = value
        ws[f'C{i}'] = unit
    
    # Calculations
    ws['A15'] = 'Load Calculations'
    ws['A16'] = 'Self Weight'
    ws['B16'] = '=B6*B7*B8'
    ws['C16'] = 'kN/m'
    
    ws['A17'] = 'Earth Pressure Coefficient Ka'
    ws['B17'] = '=TAN(RADIANS(45-B10/2))^2'
    ws['C17'] = '-'
    
    ws['A18'] = 'Active Earth Pressure'
    ws['B18'] = '=0.5*B17*B9*B6*B6'
    ws['C18'] = 'kN/m'
    
    ws['A19'] = 'Overturning Factor'
    ws['B19'] = '=B16*0.5/B18'  # Simplified
    ws['C19'] = '-'
    
    ws['A20'] = 'Safety Check'
    ws['B20'] = '=IF(B19>=2,"SAFE","CHECK")'
    
    wb.save(filename)
    print(f"✅ Created {filename}")

def create_hydraulic_file(filename, scenario):
    """Create hydraulic analysis Excel file for specific scenario"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise Exception("Failed to create worksheet")
    
    ws.title = "Hydraulic Analysis"
    
    # Headers
    ws['A1'] = f'{scenario["Bridge Name"]} - Hydraulic Analysis'
    ws['A3'] = 'Design Parameters'
    ws['A5'] = 'Parameter'
    ws['B5'] = 'Value'
    ws['C5'] = 'Unit'
    
    # Parameters based on scenario
    params = [
        ('Design Discharge', 800.0 + (scenario.get('Bridge Width', 7.5) * 20), 'cumecs'),
        ('High Flood Level', 101.2, 'm'),
        ('Bed Slope', '1 in 1000', '-'),
        ('Manning n', 0.033, '-'),
        ('Silt Factor', 1.5, '-'),
        ('Bridge Opening', scenario.get('Bridge Width', 7.5) * 10, 'm')
    ]
    
    for i, (param, value, unit) in enumerate(params, start=6):
        ws[f'A{i}'] = param
        ws[f'B{i}'] = value
        ws[f'C{i}'] = unit
    
    # Calculations
    ws['A15'] = 'Lacey Regime Calculations'
    ws['A16'] = 'Regime Width'
    ws['B16'] = '=4.75*SQRT(B6)'
    ws['C16'] = 'm'
    
    ws['A17'] = 'Flow Velocity'
    ws['B17'] = '=B6/B12'
    ws['C17'] = 'm/s'
    
    ws['A18'] = 'Adequacy Check'
    ws['B18'] = '=IF(B17<5,"ADEQUATE","REVIEW")'
    
    wb.save(filename)
    print(f"✅ Created {filename}")

def create_load_analysis_file(filename, scenario):
    """Create load analysis Excel file for specific scenario"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise Exception("Failed to create worksheet")
    
    ws.title = "Live Load Analysis"
    
    # Headers
    ws['A1'] = f'{scenario["Bridge Name"]} - Live Load Analysis'
    ws['A3'] = 'Load Parameters'
    ws['A5'] = 'Parameter'
    ws['B5'] = 'Value'
    ws['C5'] = 'Unit'
    
    # Parameters based on bridge type
    bridge_type = scenario.get("Bridge Type", "Submersible Bridge")
    if "High Level" in bridge_type:
        load_type = "Class AA (High Level)"
        load_value = 700.0
    elif "Pedestrian" in scenario.get("Bridge Name", ""):
        load_type = "Pedestrian Load"
        load_value = 5.0
    else:
        load_type = "Class A"
        load_value = 450.0
    
    params = [
        ('Load Type', load_type, '-'),
        ('Max Load', load_value, 'kN'),
        ('Load Distribution', 'Lane Load', '-'),
        ('Impact Factor', 0.25, '-'),
        ('Span Length', scenario.get('Effective Span', 10.0), 'm')
    ]
    
    for i, (param, value, unit) in enumerate(params, start=6):
        ws[f'A{i}'] = param
        # Handle text values differently
        if isinstance(value, str):
            ws[f'B{i}'] = value
        else:
            ws[f'B{i}'] = value
        ws[f'C{i}'] = unit
    
    # Calculations
    ws['A15'] = 'Load Calculations'
    ws['A16'] = 'Distributed Load'
    ws['B16'] = '=B7/B10'
    ws['C16'] = 'kN/m'
    
    ws['A17'] = 'Total Load'
    ws['B17'] = '=B16*B10'
    ws['C17'] = 'kN'
    
    ws['A18'] = 'Load Status'
    ws['B18'] = '=IF(B17<1000,"ACCEPTABLE","REVIEW")'
    
    wb.save(filename)
    print(f"✅ Created {filename}")

def create_cross_section_file(filename, scenario):
    """Create cross-section design Excel file for specific scenario"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise Exception("Failed to create worksheet")
    
    ws.title = "Cross Section Design"
    
    # Headers
    ws['A1'] = f'{scenario["Bridge Name"]} - Cross Section Design'
    ws['A3'] = 'Section Parameters'
    ws['A5'] = 'Parameter'
    ws['B5'] = 'Value'
    ws['C5'] = 'Unit'
    
    # Parameters based on scenario
    params = [
        ('Span Length', scenario.get('Effective Span', 10.0), 'm'),
        ('Section Width', scenario.get('Bridge Width', 7.5), 'm'),
        ('Slab Thickness', max(0.2, scenario.get('Effective Span', 10.0) / 20), 'm'),
        ('Concrete Grade', scenario.get('Concrete Grade', 'M25'), '-'),
        ('Steel Grade', scenario.get('Steel Grade', 'Fe415'), '-'),
        ('Design Code', scenario.get('Design Code', 'IRC-6'), '-')
    ]
    
    for i, (param, value, unit) in enumerate(params, start=6):
        ws[f'A{i}'] = param
        ws[f'B{i}'] = value
        ws[f'C{i}'] = unit
    
    # Calculations
    ws['A15'] = 'Design Calculations'
    ws['A16'] = 'Moment Capacity'
    ws['B16'] = '=B7*B8*1000'  # Simplified
    ws['C16'] = 'kN-m'
    
    ws['A17'] = 'Shear Capacity'
    ws['B17'] = '=B7*B9*500'  # Simplified
    ws['C17'] = 'kN'
    
    ws['A18'] = 'Design Status'
    ws['B18'] = '=IF(B16>1000,"APPROVED","CHECK")'
    
    wb.save(filename)
    print(f"✅ Created {filename}")

def create_abutment_file(filename, scenario):
    """Create abutment design Excel file for specific scenario"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise Exception("Failed to create worksheet")
    
    ws.title = "Abutment Design"
    
    # Headers
    ws['A1'] = f'{scenario["Bridge Name"]} - Abutment Design'
    ws['A3'] = 'Abutment Parameters'
    ws['A5'] = 'Parameter'
    ws['B5'] = 'Value'
    ws['C5'] = 'Unit'
    
    # Parameters based on scenario
    params = [
        ('Abutment Height', scenario.get('Effective Span', 10.0) * 0.7, 'm'),
        ('Backfill Angle', 30.0, 'degrees'),
        ('Coefficient of Friction', 0.5, '-'),
        ('Foundation Depth', 2.0, 'm'),
        ('Concrete Grade', scenario.get('Concrete Grade', 'M25'), '-'),
        ('Design Code', scenario.get('Design Code', 'IRC-6'), '-')
    ]
    
    for i, (param, value, unit) in enumerate(params, start=6):
        ws[f'A{i}'] = param
        ws[f'B{i}'] = value
        ws[f'C{i}'] = unit
    
    # Calculations
    ws['A15'] = 'Stability Calculations'
    ws['A16'] = 'Overturning Moment'
    ws['B16'] = '=B6*0.5*B6'  # Simplified
    ws['C16'] = 'kN-m'
    
    ws['A17'] = 'Resisting Moment'
    ws['B17'] = '=B6*B6*0.3'  # Simplified
    ws['C17'] = 'kN-m'
    
    ws['A18'] = 'Factor of Safety'
    ws['B18'] = '=B17/B16'
    ws['C18'] = '-'
    
    ws['A19'] = 'Design Status'
    ws['B19'] = '=IF(B18>2,"SAFE","REVIEW")'
    
    wb.save(filename)
    print(f"✅ Created {filename}")

def generate_test_files_for_scenario(scenario_name, scenario_data):
    """Generate all test Excel files for a specific scenario"""
    print(f"\n📁 Generating test files for: {scenario_name}")
    
    # Create test_files directory if it doesn't exist
    test_dir = Path("test_files") / scenario_name.replace(" ", "_")
    test_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate files
    files_to_create = [
        (f"{scenario_name.replace(' ', '_')}_Stability_Analysis.xlsx", create_stability_file),
        (f"{scenario_name.replace(' ', '_')}_Hydraulic_Analysis.xlsx", create_hydraulic_file),
        (f"{scenario_name.replace(' ', '_')}_Live_Load_Analysis.xlsx", create_load_analysis_file),
        (f"{scenario_name.replace(' ', '_')}_Cross_Section_Design.xlsx", create_cross_section_file),
        (f"{scenario_name.replace(' ', '_')}_Abutment_Design.xlsx", create_abutment_file)
    ]
    
    for filename, create_func in files_to_create:
        filepath = test_dir / filename
        try:
            create_func(str(filepath), scenario_data)
        except Exception as e:
            print(f"❌ Error creating {filename}: {str(e)}")

def main():
    """Generate test Excel files for all 5 test scenarios"""
    print("Creating test Excel files for all 5 test scenarios...")
    
    # Define test scenarios
    test_scenarios = {
        "Test_1_Standard_Submersible": {
            "Bridge Name": "Main Street Submersible Bridge",
            "Location": "Downtown Area",
            "Bridge Type": "Submersible Bridge",
            "Effective Span": 12.0,
            "Bridge Width": 8.5,
            "Number of Spans": 3,
            "Skew Angle": 0.0,
            "Design Code": "IRC-6",
            "Concrete Grade": "M30",
            "Steel Grade": "Fe500",
            "Design Life": 100
        },
        "Test_2_High_Level_Skew": {
            "Bridge Name": "River Crossing High-Level Bridge",
            "Location": "Northern Highway",
            "Bridge Type": "High Level Bridge",
            "Effective Span": 25.0,
            "Bridge Width": 12.0,
            "Number of Spans": 5,
            "Skew Angle": 30.0,
            "Design Code": "IRC-21",
            "Concrete Grade": "M35",
            "Steel Grade": "Fe500",
            "Design Life": 120
        },
        "Test_3_Narrow_Culvert": {
            "Bridge Name": "Forest Road Culvert",
            "Location": "Mountain Region",
            "Bridge Type": "Culvert",
            "Effective Span": 6.0,
            "Bridge Width": 4.0,
            "Number of Spans": 1,
            "Skew Angle": 15.0,
            "Design Code": "IRC-112",
            "Concrete Grade": "M25",
            "Steel Grade": "Fe415",
            "Design Life": 75
        },
        "Test_4_Wide_Aqueduct": {
            "Bridge Name": "Irrigation Aqueduct Bridge",
            "Location": "Agricultural Zone",
            "Bridge Type": "Aqueduct",
            "Effective Span": 15.0,
            "Bridge Width": 18.0,
            "Number of Spans": 4,
            "Skew Angle": 5.0,
            "Design Code": "IS-456",
            "Concrete Grade": "M40",
            "Steel Grade": "Fe550",
            "Design Life": 100
        },
        "Test_5_Urban_Pedestrian": {
            "Bridge Name": "City Park Pedestrian Bridge",
            "Location": "Urban Park",
            "Bridge Type": "Submersible Bridge",
            "Effective Span": 8.0,
            "Bridge Width": 3.5,
            "Number of Spans": 1,
            "Skew Angle": 0.0,
            "Design Code": "IRC-6",
            "Concrete Grade": "M30",
            "Steel Grade": "Fe415",
            "Design Life": 50
        }
    }
    
    # Generate files for each scenario
    for scenario_name, scenario_data in test_scenarios.items():
        generate_test_files_for_scenario(scenario_name, scenario_data)
    
    print(f"\n✅ All test Excel files generated successfully!")
    print(f"📁 Files are located in the 'test_files' directory")
    print(f"📊 Total scenarios: {len(test_scenarios)}")
    print(f"📄 Files per scenario: 5")
    print(f"🔢 Total files created: {len(test_scenarios) * 5}")

if __name__ == "__main__":
    main()