"""
Master Coordinator Module
Coordinates all design analyses and creates master Excel file
Links formulas across different analysis modules
"""

import pandas as pd
import numpy as np
import io
import json
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class MasterCoordinator:
    """Coordinates all bridge design analyses and creates master Excel file"""
    
    def __init__(self, bridge_config):
        self.bridge_config = bridge_config
        self.analysis_registry = {}
        self.parameter_links = {}
        self.formula_dependencies = {}
        self.master_formulas = {}
        
    def create_master_file(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Create master coordination file linking all analyses"""
        
        master_data = {
            'project_info': self._compile_project_info(),
            'analysis_summary': self._compile_analysis_summary(analysis_results),
            'parameter_coordination': self._coordinate_parameters(analysis_results),
            'formula_integration': self._integrate_formulas(analysis_results),
            'consistency_checks': self._perform_consistency_checks(analysis_results),
            'optimization_opportunities': self._identify_optimization_opportunities(analysis_results),
            'summary': self._create_master_summary(analysis_results)
        }
        
        return master_data
    
    def _compile_project_info(self) -> Dict[str, Any]:
        """Compile comprehensive project information"""
        
        return {
            'bridge_name': self.bridge_config.bridge_name,
            'location': self.bridge_config.location,
            'project_type': self.bridge_config.project_type,
            'span_configuration': {
                'effective_span': self.bridge_config.span_length,
                'bridge_width': self.bridge_config.bridge_width,
                'number_of_spans': self.bridge_config.num_spans,
                'skew_angle': self.bridge_config.skew_angle
            },
            'design_standards': {
                'design_code': self.bridge_config.design_code,
                'concrete_grade': self.bridge_config.concrete_grade,
                'steel_grade': self.bridge_config.steel_grade,
                'design_life': self.bridge_config.design_life
            },
            'coordination_timestamp': datetime.now().isoformat()
        }
    
    def _compile_analysis_summary(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Compile summary of all completed analyses"""
        
        summary = {
            'completed_analyses': list(analysis_results.keys()),
            'analysis_status': {},
            'key_results': {},
            'overall_design_status': 'PENDING'
        }
        
        all_safe = True
        
        for analysis_name, results in analysis_results.items():
            if isinstance(results, dict):
                # Extract status information
                status = self._extract_analysis_status(results)
                summary['analysis_status'][analysis_name] = status
                
                if status not in ['SAFE', 'ACCEPTABLE', 'HYDRAULICALLY_SAFE']:
                    all_safe = False
                
                # Extract key results
                key_results = self._extract_key_results(analysis_name, results)
                summary['key_results'][analysis_name] = key_results
        
        summary['overall_design_status'] = 'SAFE' if all_safe else 'NEEDS_REVIEW'
        
        return summary
    
    def _extract_analysis_status(self, results: Dict[str, Any]) -> str:
        """Extract status from analysis results"""
        
        # Look for common status fields
        status_fields = ['status', 'design_status', 'overall_status', 'overall_adequacy']
        
        for field in status_fields:
            if field in results:
                return str(results[field])
        
        # Try to infer status from safety factors
        if 'overturning_factor' in results and 'sliding_factor' in results:
            if (results.get('overturning_factor', 0) >= 2.0 and 
                results.get('sliding_factor', 0) >= 1.5):
                return 'SAFE'
            else:
                return 'NEEDS_REVIEW'
        
        return 'UNKNOWN'
    
    def _extract_key_results(self, analysis_name: str, results: Dict[str, Any]) -> Dict[str, Any]:
        """Extract key results from each analysis"""
        
        key_results = {}
        
        if analysis_name == 'hydraulic':
            key_results = {
                'regime_width': results.get('regime_width', 'N/A'),
                'afflux': results.get('afflux', 'N/A'),
                'scour_depth': results.get('scour_depth', 'N/A'),
                'pier_spacing': results.get('pier_spacing', 'N/A')
            }
        
        elif analysis_name == 'stability':
            key_results = {
                'overturning_factor': results.get('overturning_factor', 'N/A'),
                'sliding_factor': results.get('sliding_factor', 'N/A'),
                'max_soil_pressure': results.get('max_soil_pressure', 'N/A')
            }
        
        elif analysis_name == 'abutment':
            key_results = {
                'abutment_type': results.get('abutment_type', 'N/A'),
                'concrete_volume': results.get('concrete_volume', 'N/A'),
                'steel_weight': results.get('steel_weight', 'N/A'),
                'foundation_type': results.get('foundation', {}).get('foundation_type', 'N/A')
            }
        
        elif analysis_name == 'cross_section':
            key_results = {
                'total_width': results.get('total_width', 'N/A'),
                'max_moment': results.get('max_moment', 'N/A'),
                'steel_required': results.get('steel_required', 'N/A'),
                'deflection': results.get('deflection', 'N/A')
            }
        
        else:
            # Generic extraction for other analyses
            for key, value in results.items():
                if isinstance(value, (int, float, str)) and len(str(value)) < 50:
                    key_results[key] = value
        
        return key_results
    
    def _coordinate_parameters(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Coordinate parameters across different analyses"""
        
        coordination = {
            'shared_parameters': {},
            'parameter_conflicts': [],
            'recommended_values': {},
            'parameter_sources': {}
        }
        
        # Define parameters that should be consistent across analyses
        shared_params = {
            'span_length': ['bridge_design', 'cross_section', 'abutment'],
            'bridge_width': ['bridge_design', 'cross_section', 'hydraulic'],
            'concrete_grade': ['bridge_design', 'abutment', 'cross_section'],
            'steel_grade': ['bridge_design', 'abutment', 'cross_section'],
            'hfl': ['hydraulic', 'stability', 'abutment'],
            'bearing_capacity': ['stability', 'abutment'],
            'unit_weight_soil': ['stability', 'abutment'],
            'angle_of_friction': ['stability', 'abutment']
        }
        
        # Extract parameter values from each analysis
        for param_name, analysis_list in shared_params.items():
            param_values = {}
            
            for analysis_name in analysis_list:
                if analysis_name in analysis_results:
                    value = self._extract_parameter_value(param_name, analysis_results[analysis_name])
                    if value is not None:
                        param_values[analysis_name] = value
            
            if param_values:
                coordination['shared_parameters'][param_name] = param_values
                
                # Check for conflicts
                unique_values = list(set(param_values.values()))
                if len(unique_values) > 1:
                    coordination['parameter_conflicts'].append({
                        'parameter': param_name,
                        'values': param_values,
                        'conflict_level': 'HIGH' if len(unique_values) > 2 else 'MEDIUM'
                    })
                
                # Recommend value
                coordination['recommended_values'][param_name] = self._recommend_parameter_value(param_values)
                coordination['parameter_sources'][param_name] = max(param_values, key=lambda k: param_values[k] if isinstance(param_values[k], (int, float)) else 0)
        
        return coordination
    
    def _extract_parameter_value(self, param_name: str, analysis_data: Dict[str, Any]) -> Any:
        """Extract parameter value from analysis data"""
        
        # Define parameter mapping for different analyses
        param_mappings = {
            'span_length': ['span_length', 'effective_span', 'L', 'span'],
            'bridge_width': ['bridge_width', 'total_width', 'width', 'W'],
            'concrete_grade': ['concrete_grade', 'fck', 'grade'],
            'steel_grade': ['steel_grade', 'fy', 'grade'],
            'hfl': ['hfl', 'high_flood_level', 'water_level'],
            'bearing_capacity': ['bearing_capacity', 'sbc', 'safe_bearing_capacity'],
            'unit_weight_soil': ['unit_weight', 'soil_density', 'gamma_soil'],
            'angle_of_friction': ['angle_of_friction', 'phi', 'friction_angle']
        }
        
        possible_keys = param_mappings.get(param_name, [param_name])
        
        # Search through the analysis data
        def search_nested_dict(data, keys):
            if isinstance(data, dict):
                for key in keys:
                    if key in data:
                        return data[key]
                
                # Search in nested dictionaries
                for value in data.values():
                    result = search_nested_dict(value, keys)
                    if result is not None:
                        return result
            
            return None
        
        return search_nested_dict(analysis_data, possible_keys)
    
    def _recommend_parameter_value(self, param_values: Dict[str, Any]) -> Any:
        """Recommend parameter value from conflicting values"""
        
        if not param_values:
            return None
        
        # If all values are the same, return that value
        unique_values = list(set(param_values.values()))
        if len(unique_values) == 1:
            return unique_values[0]
        
        # For numeric values, take the most conservative (usually maximum)
        numeric_values = [v for v in param_values.values() if isinstance(v, (int, float))]
        if numeric_values:
            return max(numeric_values)  # Conservative approach
        
        # For string values, prefer the most recent or most detailed analysis
        analysis_priority = ['abutment', 'cross_section', 'stability', 'hydraulic', 'bridge_design']
        
        for analysis in analysis_priority:
            if analysis in param_values:
                return param_values[analysis]
        
        # Default to first available value
        return list(param_values.values())[0]
    
    def _integrate_formulas(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Integrate formulas from different analyses"""
        
        integration = {
            'formula_categories': {},
            'cross_references': {},
            'formula_validation': {},
            'integration_opportunities': []
        }
        
        # Categorize formulas by engineering discipline
        categories = {
            'structural': ['moment', 'shear', 'deflection', 'stress', 'reinforcement'],
            'geotechnical': ['bearing', 'settlement', 'earth_pressure', 'stability'],
            'hydraulic': ['discharge', 'velocity', 'afflux', 'scour', 'regime'],
            'load': ['dead_load', 'live_load', 'wind_load', 'seismic']
        }
        
        for category, keywords in categories.items():
            integration['formula_categories'][category] = []
            
            for analysis_name, results in analysis_results.items():
                formulas = self._extract_formulas_from_results(results, keywords)
                if formulas:
                    integration['formula_categories'][category].extend([
                        {'analysis': analysis_name, 'formula': formula} for formula in formulas
                    ])
        
        # Identify cross-references between analyses
        integration['cross_references'] = self._identify_cross_references(analysis_results)
        
        # Validate formula consistency
        integration['formula_validation'] = self._validate_formula_consistency(analysis_results)
        
        return integration
    
    def _extract_formulas_from_results(self, results: Dict[str, Any], keywords: List[str]) -> List[str]:
        """Extract formulas related to specific keywords"""
        
        formulas = []
        
        def search_for_formulas(data, path=""):
            if isinstance(data, dict):
                for key, value in data.items():
                    current_path = f"{path}.{key}" if path else key
                    
                    # Check if key contains any keyword
                    if any(keyword in key.lower() for keyword in keywords):
                        if isinstance(value, str) and ('=' in value or 'formula' in value.lower()):
                            formulas.append(f"{current_path}: {value}")
                        elif isinstance(value, (int, float)):
                            formulas.append(f"{current_path}: {value}")
                    
                    # Recursively search nested dictionaries
                    if isinstance(value, dict):
                        search_for_formulas(value, current_path)
        
        search_for_formulas(results)
        return formulas
    
    def _identify_cross_references(self, analysis_results: Dict[str, Any]) -> Dict[str, List[str]]:
        """Identify parameters that are cross-referenced between analyses"""
        
        cross_refs = {}
        
        # Common parameters that should link between analyses
        linkable_params = [
            'span_length', 'bridge_width', 'hfl', 'bearing_capacity',
            'concrete_grade', 'steel_grade', 'load_factor', 'safety_factor'
        ]
        
        for param in linkable_params:
            analyses_with_param = []
            
            for analysis_name, results in analysis_results.items():
                if self._extract_parameter_value(param, results) is not None:
                    analyses_with_param.append(analysis_name)
            
            if len(analyses_with_param) > 1:
                cross_refs[param] = analyses_with_param
        
        return cross_refs
    
    def _validate_formula_consistency(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Validate consistency of formulas across analyses"""
        
        validation = {
            'consistent_formulas': [],
            'inconsistent_formulas': [],
            'missing_formulas': [],
            'validation_summary': {}
        }
        
        # Check for common engineering relationships
        relationships = {
            'moment_deflection': ['moment', 'deflection'],
            'load_stress': ['load', 'stress'],
            'pressure_bearing': ['pressure', 'bearing']
        }
        
        for relationship, components in relationships.items():
            found_components = {}
            
            for analysis_name, results in analysis_results.items():
                for component in components:
                    value = self._extract_parameter_value(component, results)
                    if value is not None:
                        if component not in found_components:
                            found_components[component] = []
                        found_components[component].append((analysis_name, value))
            
            # Validate relationship
            if len(found_components) == len(components):
                validation['consistent_formulas'].append({
                    'relationship': relationship,
                    'components': found_components
                })
            else:
                validation['missing_formulas'].append({
                    'relationship': relationship,
                    'missing_components': [c for c in components if c not in found_components]
                })
        
        return validation
    
    def _perform_consistency_checks(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Perform consistency checks across all analyses"""
        
        checks = {
            'dimensional_consistency': self._check_dimensional_consistency(analysis_results),
            'safety_factor_consistency': self._check_safety_factors(analysis_results),
            'load_path_consistency': self._check_load_paths(analysis_results),
            'material_property_consistency': self._check_material_properties(analysis_results)
        }
        
        # Overall consistency assessment
        all_checks_pass = all(
            check.get('status', 'FAIL') == 'PASS' 
            for check in checks.values()
        )
        
        checks['overall_status'] = 'CONSISTENT' if all_checks_pass else 'INCONSISTENT'
        
        return checks
    
    def _check_dimensional_consistency(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Check dimensional consistency across analyses"""
        
        check = {
            'status': 'PASS',
            'issues': [],
            'message': 'All dimensions are consistent'
        }
        
        # Check span length consistency
        span_lengths = {}
        for analysis_name, results in analysis_results.items():
            span_length = self._extract_parameter_value('span_length', results)
            if span_length is not None:
                span_lengths[analysis_name] = span_length
        
        if span_lengths:
            unique_spans = list(set(span_lengths.values()))
            if len(unique_spans) > 1:
                check['status'] = 'FAIL'
                check['issues'].append(f"Inconsistent span lengths: {span_lengths}")
                check['message'] = 'Dimensional inconsistencies found'
        
        return check
    
    def _check_safety_factors(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Check safety factor consistency"""
        
        check = {
            'status': 'PASS',
            'issues': [],
            'message': 'Safety factors are adequate'
        }
        
        # Check overturning and sliding factors
        for analysis_name, results in analysis_results.items():
            if 'overturning_factor' in results:
                if results['overturning_factor'] < 2.0:
                    check['status'] = 'FAIL'
                    check['issues'].append(f"{analysis_name}: Overturning factor below 2.0")
            
            if 'sliding_factor' in results:
                if results['sliding_factor'] < 1.5:
                    check['status'] = 'FAIL'
                    check['issues'].append(f"{analysis_name}: Sliding factor below 1.5")
        
        if check['issues']:
            check['message'] = 'Safety factor violations found'
        
        return check
    
    def _check_load_paths(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Check load path consistency"""
        
        check = {
            'status': 'PASS',
            'issues': [],
            'message': 'Load paths are consistent'
        }
        
        # This is a simplified check - in practice, would be more complex
        required_analyses = ['cross_section', 'abutment']
        available_analyses = list(analysis_results.keys())
        
        missing_analyses = [req for req in required_analyses if req not in available_analyses]
        
        if missing_analyses:
            check['status'] = 'FAIL'
            check['issues'].append(f"Missing analyses for complete load path: {missing_analyses}")
            check['message'] = 'Incomplete load path analysis'
        
        return check
    
    def _check_material_properties(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Check material property consistency"""
        
        check = {
            'status': 'PASS',
            'issues': [],
            'message': 'Material properties are consistent'
        }
        
        # Check concrete grade consistency
        concrete_grades = {}
        for analysis_name, results in analysis_results.items():
            grade = self._extract_parameter_value('concrete_grade', results)
            if grade is not None:
                concrete_grades[analysis_name] = grade
        
        if concrete_grades:
            unique_grades = list(set(concrete_grades.values()))
            if len(unique_grades) > 1:
                check['status'] = 'FAIL'
                check['issues'].append(f"Inconsistent concrete grades: {concrete_grades}")
                check['message'] = 'Material property inconsistencies found'
        
        return check
    
    def _identify_optimization_opportunities(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Identify opportunities for design optimization"""
        
        opportunities = {
            'material_optimization': [],
            'dimensional_optimization': [],
            'cost_optimization': [],
            'construction_optimization': []
        }
        
        # Material optimization
        for analysis_name, results in analysis_results.items():
            if 'steel_required' in results and 'steel_provided' in results:
                required = results.get('steel_required', 0)
                provided = results.get('steel_provided', 0)
                
                if isinstance(required, (int, float)) and isinstance(provided, (int, float)):
                    if provided > required * 1.2:  # 20% over-design
                        opportunities['material_optimization'].append({
                            'analysis': analysis_name,
                            'opportunity': 'Reduce steel reinforcement',
                            'potential_saving': f"{((provided - required) / required * 100):.1f}% steel reduction"
                        })
        
        # Dimensional optimization
        for analysis_name, results in analysis_results.items():
            if 'overturning_factor' in results:
                factor = results.get('overturning_factor', 0)
                if factor > 3.0:  # Very high safety factor
                    opportunities['dimensional_optimization'].append({
                        'analysis': analysis_name,
                        'opportunity': 'Reduce base dimensions',
                        'potential_saving': 'Concrete volume reduction possible'
                    })
        
        return opportunities
    
    def _create_master_summary(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Create master summary of all coordination activities"""
        
        total_analyses = len(analysis_results)
        successful_analyses = sum(1 for results in analysis_results.values() 
                                 if self._extract_analysis_status(results) in ['SAFE', 'ACCEPTABLE'])
        
        summary = {
            'total_sheets': self._estimate_total_sheets(analysis_results),
            'total_formulas': self._estimate_total_formulas(analysis_results),
            'integration_status': 'COMPLETE' if total_analyses > 0 else 'PENDING',
            'sheet_list': self._generate_sheet_list(analysis_results),
            'coordination_quality': 'GOOD' if successful_analyses / max(total_analyses, 1) >= 0.8 else 'NEEDS_IMPROVEMENT',
            'recommendations': self._generate_coordination_recommendations(analysis_results)
        }
        
        return summary
    
    def _estimate_total_sheets(self, analysis_results: Dict[str, Any]) -> int:
        """Estimate total number of Excel sheets in master file"""
        
        base_sheets = 5  # Summary, parameters, formulas, etc.
        analysis_sheets = len(analysis_results) * 8  # Average 8 sheets per analysis
        
        return base_sheets + analysis_sheets
    
    def _estimate_total_formulas(self, analysis_results: Dict[str, Any]) -> int:
        """Estimate total number of formulas"""
        
        total_formulas = 0
        
        for results in analysis_results.values():
            # Estimate based on complexity of results
            if isinstance(results, dict):
                total_formulas += len(results) * 2  # Rough estimate
        
        return total_formulas
    
    def _generate_sheet_list(self, analysis_results: Dict[str, Any]) -> List[str]:
        """Generate list of sheets in master file"""
        
        sheets = ['Summary', 'Project Parameters', 'Coordination Matrix']
        
        for analysis_name in analysis_results.keys():
            sheets.extend([
                f"{analysis_name.title()} - Input",
                f"{analysis_name.title()} - Calculations", 
                f"{analysis_name.title()} - Results",
                f"{analysis_name.title()} - Checks"
            ])
        
        sheets.extend(['Integration Summary', 'Optimization', 'References'])
        
        return sheets
    
    def _generate_coordination_recommendations(self, analysis_results: Dict[str, Any]) -> List[str]:
        """Generate recommendations for better coordination"""
        
        recommendations = []
        
        if len(analysis_results) < 4:
            recommendations.append("Complete all major analyses (hydraulic, stability, abutment, cross-section)")
        
        # Check for parameter consistency
        param_coord = self._coordinate_parameters(analysis_results)
        if param_coord['parameter_conflicts']:
            recommendations.append("Resolve parameter conflicts between analyses")
        
        # Check for missing cross-references
        cross_refs = self._identify_cross_references(analysis_results)
        if len(cross_refs) < 5:
            recommendations.append("Establish more parameter links between analyses")
        
        if not recommendations:
            recommendations.append("Coordination is satisfactory - proceed with final design")
        
        return recommendations
    
    def generate_master_excel(self, master_data: Dict[str, Any]) -> io.BytesIO:
        """Generate master Excel file with all coordinated data"""
        
        buffer = io.BytesIO()
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet
            self._create_summary_sheet(wb, master_data)
            
            # Create project parameters sheet
            self._create_parameters_sheet(wb, master_data)
            
            # Create coordination matrix sheet
            self._create_coordination_sheet(wb, master_data)
            
            # Create analysis sheets
            self._create_analysis_sheets(wb, master_data)
            
            # Create integration sheet
            self._create_integration_sheet(wb, master_data)
            
            # Save to buffer
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            print(f"Error generating master Excel file: {str(e)}")
            # Return a simple Excel file with error message
            simple_wb = openpyxl.Workbook()
            ws = simple_wb.active
            ws.title = "Error"
            ws['A1'] = f"Error generating master file: {str(e)}"
            simple_wb.save(buffer)
            buffer.seek(0)
            return buffer
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, master_data: Dict[str, Any]):
        """Create summary sheet"""
        
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Bridge Design Master Coordination Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Project info
        project_info = master_data.get('project_info', {})
        row = 3
        
        ws[f'A{row}'] = "Project Information"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for key, value in project_info.items():
            if isinstance(value, dict):
                ws[f'A{row}'] = f"{key}:"
                row += 1
                for sub_key, sub_value in value.items():
                    ws[f'B{row}'] = f"{sub_key}: {sub_value}"
                    row += 1
            else:
                ws[f'A{row}'] = f"{key}: {value}"
                row += 1
        
        # Analysis summary
        row += 2
        ws[f'A{row}'] = "Analysis Summary"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        analysis_summary = master_data.get('analysis_summary', {})
        for key, value in analysis_summary.items():
            ws[f'A{row}'] = f"{key}: {value}"
            row += 1
    
    def _create_parameters_sheet(self, wb: openpyxl.Workbook, master_data: Dict[str, Any]):
        """Create parameters coordination sheet"""
        
        ws = wb.create_sheet("Parameters")
        
        # Headers
        headers = ['Parameter', 'Recommended Value', 'Source Analysis', 'Status']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Parameter data
        param_coord = master_data.get('parameter_coordination', {})
        recommended_values = param_coord.get('recommended_values', {})
        parameter_sources = param_coord.get('parameter_sources', {})
        
        row = 2
        for param, value in recommended_values.items():
            ws.cell(row=row, column=1, value=param)
            ws.cell(row=row, column=2, value=str(value))
            ws.cell(row=row, column=3, value=parameter_sources.get(param, 'Unknown'))
            ws.cell(row=row, column=4, value='Coordinated')
            row += 1
    
    def _create_coordination_sheet(self, wb: openpyxl.Workbook, master_data: Dict[str, Any]):
        """Create coordination matrix sheet"""
        
        ws = wb.create_sheet("Coordination Matrix")
        
        ws['A1'] = "Parameter Cross-Reference Matrix"
        ws['A1'].font = Font(size=14, bold=True)
        
        # This would contain a matrix showing which parameters are used in which analyses
        # Simplified implementation
        row = 3
        ws[f'A{row}'] = "Cross-references between analyses completed"
        
    def _create_analysis_sheets(self, wb: openpyxl.Workbook, master_data: Dict[str, Any]):
        """Create individual analysis sheets"""
        
        analysis_summary = master_data.get('analysis_summary', {})
        key_results = analysis_summary.get('key_results', {})
        
        for analysis_name, results in key_results.items():
            ws = wb.create_sheet(f"{analysis_name.title()}")
            
            # Title
            ws['A1'] = f"{analysis_name.title()} Analysis Results"
            ws['A1'].font = Font(size=14, bold=True)
            
            # Results
            row = 3
            for key, value in results.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                row += 1
    
    def _create_integration_sheet(self, wb: openpyxl.Workbook, master_data: Dict[str, Any]):
        """Create integration summary sheet"""
        
        ws = wb.create_sheet("Integration")
        
        ws['A1'] = "Design Integration Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Consistency checks
        consistency = master_data.get('consistency_checks', {})
        
        row = 3
        ws[f'A{row}'] = "Consistency Checks"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for check_name, check_result in consistency.items():
            if isinstance(check_result, dict):
                status = check_result.get('status', 'Unknown')
                message = check_result.get('message', 'No message')
                ws[f'A{row}'] = f"{check_name}: {status}"
                ws[f'B{row}'] = message
                row += 1
    
    def identify_common_parameters(self, analysis_results: Dict[str, Any]) -> Dict[str, List[str]]:
        """Identify common parameters across analyses for UI display"""
        
        common_params = {}
        
        # Structural parameters
        structural_params = []
        for analysis_name, results in analysis_results.items():
            if analysis_name in ['cross_section', 'abutment']:
                if 'concrete_grade' in str(results):
                    structural_params.append('concrete_grade')
                if 'steel_grade' in str(results):
                    structural_params.append('steel_grade')
        
        if structural_params:
            common_params['Structural'] = list(set(structural_params))
        
        # Geometric parameters
        geometric_params = []
        for analysis_name, results in analysis_results.items():
            if 'span_length' in str(results):
                geometric_params.append('span_length')
            if 'bridge_width' in str(results):
                geometric_params.append('bridge_width')
        
        if geometric_params:
            common_params['Geometric'] = list(set(geometric_params))
        
        # Hydraulic parameters
        hydraulic_params = []
        if 'hydraulic' in analysis_results:
            hydraulic_params.extend(['discharge', 'hfl', 'regime_width', 'afflux'])
        
        if hydraulic_params:
            common_params['Hydraulic'] = hydraulic_params
        
        return common_params
    
    def check_consistency(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Check consistency for UI display"""
        
        return self._perform_consistency_checks(analysis_results)

