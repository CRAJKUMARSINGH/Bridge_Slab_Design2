"""
Excel File Processor Module
Extracts formulas, data, and maintains original calculation logic from PROJECT FILES
"""

import pandas as pd
import openpyxl
from openpyxl.formula.translate import Translator
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
import re
import io
import json
from datetime import datetime

class ExcelProcessor:
    """Processes Excel files from Bridge_Slab_Design PROJECT FILES"""
    
    def __init__(self):
        self.supported_extensions = ['.xls', '.xlsx']
        self.formula_cache = {}
        self.data_cache = {}
    
    def process_excel_file(self, file_buffer) -> Dict[str, Any]:
        """
        Process uploaded Excel file and extract all formulas and data
        Maintains original formulas from Bundan River Bridge TAD project structure
        """
        try:
            # Read the Excel file
            if hasattr(file_buffer, 'read'):
                file_buffer.seek(0)
                wb = openpyxl.load_workbook(file_buffer, data_only=False)
            else:
                wb = openpyxl.load_workbook(file_buffer, data_only=False)
            
            processed_data = {
                'filename': getattr(file_buffer, 'name', 'unknown.xlsx'),
                'sheets': {},
                'metadata': {
                    'processed_date': datetime.now().isoformat(),
                    'total_sheets': len(wb.sheetnames),
                    'sheet_names': wb.sheetnames
                }
            }
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet_data = self._process_sheet(wb[sheet_name])
                processed_data['sheets'][sheet_name] = sheet_data
            
            return processed_data
            
        except Exception as e:
            raise Exception(f"Error processing Excel file: {str(e)}")
    
    def _process_sheet(self, worksheet) -> Dict[str, Any]:
        """Process individual worksheet and extract formulas and data"""
        
        sheet_data = {
            'formulas': {},
            'values': {},
            'data_ranges': {},
            'charts': [],
            'metadata': {
                'max_row': worksheet.max_row,
                'max_column': worksheet.max_column,
                'sheet_name': worksheet.title
            }
        }
        
        # Extract formulas and values
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_ref = f"{cell.column_letter}{cell.row}"
                    
                    # Store cell value
                    sheet_data['values'][cell_ref] = cell.value
                    
                    # Extract formula if present
                    if hasattr(cell, 'formula') and cell.formula:
                        sheet_data['formulas'][cell_ref] = cell.formula
                    elif isinstance(cell.value, str) and cell.value.startswith('='):
                        sheet_data['formulas'][cell_ref] = cell.value
        
        # Identify data ranges (tables, named ranges, etc.)
        sheet_data['data_ranges'] = self._identify_data_ranges(worksheet)
        
        # Extract chart information
        if hasattr(worksheet, '_charts'):
            for chart in worksheet._charts:
                chart_info = {
                    'type': type(chart).__name__,
                    'title': getattr(chart, 'title', 'Unknown'),
                    'anchor': str(chart.anchor) if hasattr(chart, 'anchor') else None
                }
                sheet_data['charts'].append(chart_info)
        
        return sheet_data
    
    def _identify_data_ranges(self, worksheet) -> Dict[str, Any]:
        """Identify important data ranges in the worksheet"""
        
        ranges = {}
        
        # Look for typical bridge design data patterns
        patterns = {
            'stability_data': [
                'overturning', 'moment', 'factor', 'safety',
                'vertical', 'horizontal', 'load', 'pressure'
            ],
            'hydraulic_data': [
                'discharge', 'velocity', 'afflux', 'scour',
                'hfl', 'water', 'flow', 'regime'
            ],
            'material_data': [
                'concrete', 'steel', 'grade', 'strength',
                'modulus', 'density', 'yield'
            ],
            'geometry_data': [
                'span', 'width', 'thickness', 'height',
                'length', 'dimension', 'spacing'
            ]
        }
        
        for range_type, keywords in patterns.items():
            range_cells = []
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.lower()
                        if any(keyword in cell_text for keyword in keywords):
                            range_cells.append(f"{cell.column_letter}{cell.row}")
            
            if range_cells:
                ranges[range_type] = {
                    'cells': range_cells,
                    'count': len(range_cells)
                }
        
        return ranges
    
    def extract_calculation_logic(self, processed_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract calculation logic and formula dependencies
        Based on actual PROJECT FILES structure
        """
        
        calculation_logic = {
            'formula_dependencies': {},
            'calculation_sequences': {},
            'parameter_mappings': {},
            'validation_rules': {}
        }
        
        # Analyze formula dependencies
        for sheet_name, sheet_data in processed_data['sheets'].items():
            if 'formulas' in sheet_data:
                dependencies = self._analyze_formula_dependencies(sheet_data['formulas'])
                calculation_logic['formula_dependencies'][sheet_name] = dependencies
        
        # Identify calculation sequences
        calculation_logic['calculation_sequences'] = self._identify_calculation_sequences(processed_data)
        
        # Map parameters to their physical meanings
        calculation_logic['parameter_mappings'] = self._map_parameters_to_meanings(processed_data)
        
        return calculation_logic
    
    def _analyze_formula_dependencies(self, formulas: Dict[str, str]) -> Dict[str, List[str]]:
        """Analyze dependencies between formulas"""
        
        dependencies = {}
        
        for cell_ref, formula in formulas.items():
            # Extract cell references from formula
            cell_pattern = r'[A-Z]+\d+'
            referenced_cells = re.findall(cell_pattern, formula)
            
            # Remove the current cell from references
            referenced_cells = [ref for ref in referenced_cells if ref != cell_ref]
            
            if referenced_cells:
                dependencies[cell_ref] = list(set(referenced_cells))
        
        return dependencies
    
    def _identify_calculation_sequences(self, processed_data: Dict[str, Any]) -> Dict[str, List[str]]:
        """Identify logical calculation sequences"""
        
        sequences = {}
        
        # Common bridge design calculation sequences
        sequence_patterns = {
            'stability_analysis': [
                'dead_load_calculation',
                'earth_pressure_calculation', 
                'moment_calculation',
                'overturning_check',
                'sliding_check',
                'bearing_pressure_check'
            ],
            'hydraulic_analysis': [
                'regime_width_calculation',
                'velocity_calculation',
                'afflux_calculation',
                'scour_depth_calculation',
                'waterway_adequacy_check'
            ],
            'structural_design': [
                'load_calculation',
                'moment_calculation',
                'reinforcement_design',
                'deflection_check',
                'crack_width_check'
            ]
        }
        
        # Match formulas to sequences based on content
        for sequence_name, steps in sequence_patterns.items():
            sequence_formulas = []
            
            for sheet_name, sheet_data in processed_data['sheets'].items():
                if 'formulas' in sheet_data:
                    for cell_ref, formula in sheet_data['formulas'].items():
                        # Simple keyword matching - can be enhanced
                        formula_lower = formula.lower()
                        for step in steps:
                            step_keywords = step.replace('_', ' ').split()
                            if any(keyword in formula_lower for keyword in step_keywords):
                                sequence_formulas.append({
                                    'step': step,
                                    'sheet': sheet_name,
                                    'cell': cell_ref,
                                    'formula': formula
                                })
                                break
            
            if sequence_formulas:
                sequences[sequence_name] = sequence_formulas
        
        return sequences
    
    def _map_parameters_to_meanings(self, processed_data: Dict[str, Any]) -> Dict[str, str]:
        """Map cell references to their physical engineering meanings"""
        
        parameter_mappings = {}
        
        # Common bridge engineering parameter patterns
        parameter_patterns = {
            'L': 'Span Length',
            'W': 'Width', 
            'H': 'Height',
            'D': 'Depth',
            'Q': 'Discharge',
            'V': 'Velocity',
            'M': 'Moment',
            'P': 'Load/Pressure',
            'F': 'Force',
            'γ': 'Unit Weight',
            'φ': 'Angle of Friction',
            'f': 'Stress/Strength',
            'E': 'Modulus',
            'ν': 'Poisson Ratio',
            'σ': 'Stress',
            'τ': 'Shear Stress'
        }
        
        # Enhanced mapping based on context
        context_mappings = {
            'span': 'Effective Span Length (m)',
            'width': 'Bridge Width (m)',
            'discharge': 'Design Discharge (cumecs)',
            'velocity': 'Design Velocity (m/s)',
            'hfl': 'High Flood Level (m)',
            'moment': 'Bending Moment (kN-m)',
            'pressure': 'Soil Pressure (kN/m²)',
            'factor': 'Safety Factor',
            'grade': 'Material Grade',
            'thickness': 'Section Thickness (m)',
            'reinforcement': 'Steel Reinforcement (mm²)',
            'concrete': 'Concrete Grade/Strength',
            'load': 'Applied Load (kN)',
            'deflection': 'Deflection (mm)',
            'stress': 'Stress (N/mm²)'
        }
        
        # Analyze cell contents and nearby text to determine meanings
        for sheet_name, sheet_data in processed_data['sheets'].items():
            if 'values' in sheet_data:
                for cell_ref, value in sheet_data['values'].items():
                    if isinstance(value, str):
                        value_lower = value.lower()
                        
                        # Check for direct matches
                        for keyword, meaning in context_mappings.items():
                            if keyword in value_lower:
                                parameter_mappings[f"{sheet_name}!{cell_ref}"] = meaning
                                break
                        
                        # Check for symbol patterns
                        for symbol, meaning in parameter_patterns.items():
                            if symbol in value:
                                parameter_mappings[f"{sheet_name}!{cell_ref}"] = meaning
                                break
        
        return parameter_mappings
    
    def validate_formulas(self, processed_data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate extracted formulas for consistency and correctness"""
        
        validation_results = {
            'syntax_errors': [],
            'reference_errors': [],
            'circular_references': [],
            'warnings': [],
            'statistics': {}
        }
        
        total_formulas = 0
        valid_formulas = 0
        
        for sheet_name, sheet_data in processed_data['sheets'].items():
            if 'formulas' in sheet_data:
                for cell_ref, formula in sheet_data['formulas'].items():
                    total_formulas += 1
                    
                    # Basic syntax validation
                    if self._validate_formula_syntax(formula):
                        valid_formulas += 1
                    else:
                        validation_results['syntax_errors'].append({
                            'sheet': sheet_name,
                            'cell': cell_ref,
                            'formula': formula,
                            'error': 'Invalid syntax'
                        })
                    
                    # Check for circular references
                    if self._check_circular_reference(cell_ref, formula, sheet_data['formulas']):
                        validation_results['circular_references'].append({
                            'sheet': sheet_name,
                            'cell': cell_ref,
                            'formula': formula
                        })
        
        validation_results['statistics'] = {
            'total_formulas': total_formulas,
            'valid_formulas': valid_formulas,
            'validation_rate': valid_formulas / max(total_formulas, 1) * 100
        }
        
        return validation_results
    
    def _validate_formula_syntax(self, formula: str) -> bool:
        """Basic formula syntax validation"""
        try:
            # Check for balanced parentheses
            if formula.count('(') != formula.count(')'):
                return False
            
            # Check for valid Excel formula start
            if not formula.startswith('='):
                return False
            
            # Check for invalid characters
            invalid_chars = ['#', '@', '&', '|']
            if any(char in formula for char in invalid_chars):
                return False
            
            return True
            
        except Exception:
            return False
    
    def _check_circular_reference(self, cell_ref: str, formula: str, all_formulas: Dict[str, str]) -> bool:
        """Check for circular references in formulas"""
        
        def get_references(formula_text: str) -> List[str]:
            cell_pattern = r'[A-Z]+\d+'
            return re.findall(cell_pattern, formula_text)
        
        def has_circular_ref(current_cell: str, target_cell: str, formulas: Dict[str, str], visited: set) -> bool:
            if current_cell in visited:
                return True
            
            if current_cell not in formulas:
                return False
            
            visited.add(current_cell)
            
            references = get_references(formulas[current_cell])
            
            for ref in references:
                if ref == target_cell:
                    return True
                if has_circular_ref(ref, target_cell, formulas, visited.copy()):
                    return True
            
            return False
        
        references = get_references(formula)
        
        for ref in references:
            if has_circular_ref(ref, cell_ref, all_formulas, set()):
                return True
        
        return False
    
    def export_formulas_to_json(self, processed_data: Dict[str, Any], filename: str = None) -> str:
        """Export extracted formulas to JSON format for integration"""
        
        export_data = {
            'metadata': {
                'export_date': datetime.now().isoformat(),
                'source_file': processed_data.get('filename', 'unknown'),
                'total_sheets': len(processed_data.get('sheets', {}))
            },
            'formulas': {},
            'calculation_logic': self.extract_calculation_logic(processed_data),
            'validation_results': self.validate_formulas(processed_data)
        }
        
        # Flatten formulas with sheet prefixes
        for sheet_name, sheet_data in processed_data.get('sheets', {}).items():
            if 'formulas' in sheet_data:
                for cell_ref, formula in sheet_data['formulas'].items():
                    full_ref = f"{sheet_name}!{cell_ref}"
                    export_data['formulas'][full_ref] = formula
        
        json_output = json.dumps(export_data, indent=2, ensure_ascii=False)
        
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(json_output)
        
        return json_output
    
    def create_master_formula_mapping(self, multiple_files: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Create master mapping of formulas across multiple Excel files"""
        
        master_mapping = {
            'global_formulas': {},
            'cross_file_dependencies': {},
            'parameter_conflicts': {},
            'integration_points': {},
            'metadata': {
                'files_processed': len(multiple_files),
                'creation_date': datetime.now().isoformat()
            }
        }
        
        # Collect all formulas
        for file_data in multiple_files:
            filename = file_data.get('filename', 'unknown')
            
            for sheet_name, sheet_data in file_data.get('sheets', {}).items():
                if 'formulas' in sheet_data:
                    for cell_ref, formula in sheet_data['formulas'].items():
                        global_ref = f"{filename}#{sheet_name}!{cell_ref}"
                        master_mapping['global_formulas'][global_ref] = {
                            'formula': formula,
                            'source_file': filename,
                            'sheet': sheet_name,
                            'cell': cell_ref
                        }
        
        # Identify integration points
        master_mapping['integration_points'] = self._identify_integration_points(multiple_files)
        
        return master_mapping
    
    def _identify_integration_points(self, multiple_files: List[Dict[str, Any]]) -> Dict[str, List[str]]:
        """Identify common parameters across files for integration"""
        
        integration_points = {}
        
        # Common bridge design integration parameters
        common_parameters = [
            'span_length', 'bridge_width', 'design_load',
            'concrete_grade', 'steel_grade', 'hfl',
            'discharge', 'bearing_capacity', 'unit_weight'
        ]
        
        for param in common_parameters:
            param_locations = []
            
            for file_data in multiple_files:
                filename = file_data.get('filename', 'unknown')
                
                for sheet_name, sheet_data in file_data.get('sheets', {}).items():
                    # Look for parameter in values and formulas
                    for cell_ref, value in sheet_data.get('values', {}).items():
                        if isinstance(value, str) and param.replace('_', ' ') in value.lower():
                            param_locations.append(f"{filename}#{sheet_name}!{cell_ref}")
            
            if len(param_locations) > 1:  # Found in multiple locations
                integration_points[param] = param_locations
        
        return integration_points
