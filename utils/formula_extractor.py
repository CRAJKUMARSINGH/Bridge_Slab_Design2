"""
Formula Extractor Utility
Extracts and processes formulas from Excel files maintaining original structure
"""

import re
import numpy as np
import pandas as pd
from typing import Dict, List, Any, Optional, Tuple, Union
import openpyxl
from openpyxl.formula.translate import Translator
import ast
import operator

class FormulaExtractor:
    """Extract and process formulas from Excel files"""
    
    def __init__(self):
        self.operators = {
            ast.Add: operator.add,
            ast.Sub: operator.sub,
            ast.Mult: operator.mul,
            ast.Div: operator.truediv,
            ast.Pow: operator.pow,
            ast.USub: operator.neg,
        }
        
        self.excel_functions = {
            'SUM': np.sum,
            'AVERAGE': np.mean,
            'MAX': np.max,
            'MIN': np.min,
            'SQRT': np.sqrt,
            'ABS': abs,
            'ROUND': round,
            'IF': self._excel_if,
            'AND': self._excel_and,
            'OR': self._excel_or,
            'POWER': pow,
            'EXP': np.exp,
            'LN': np.log,
            'LOG': np.log10,
            'SIN': np.sin,
            'COS': np.cos,
            'TAN': np.tan,
            'PI': lambda: np.pi
        }
    
    def extract_formulas_from_workbook(self, workbook_path: str) -> Dict[str, Any]:
        """Extract all formulas from an Excel workbook"""
        
        try:
            wb = openpyxl.load_workbook(workbook_path, data_only=False)
            
            extracted_data = {
                'workbook_name': workbook_path,
                'sheets': {},
                'formula_summary': {
                    'total_formulas': 0,
                    'formula_types': {},
                    'dependencies': {},
                    'circular_references': []
                }
            }
            
            for sheet_name in wb.sheetnames:
                sheet_data = self._extract_sheet_formulas(wb[sheet_name])
                extracted_data['sheets'][sheet_name] = sheet_data
                extracted_data['formula_summary']['total_formulas'] += len(sheet_data.get('formulas', {}))
            
            # Analyze formula dependencies across sheets
            extracted_data['formula_summary']['dependencies'] = self._analyze_cross_sheet_dependencies(extracted_data['sheets'])
            
            return extracted_data
            
        except Exception as e:
            raise Exception(f"Error extracting formulas from {workbook_path}: {str(e)}")
    
    def _extract_sheet_formulas(self, worksheet) -> Dict[str, Any]:
        """Extract formulas from a single worksheet"""
        
        sheet_data = {
            'sheet_name': worksheet.title,
            'formulas': {},
            'values': {},
            'formula_metadata': {},
            'named_ranges': {},
            'data_validation': {}
        }
        
        # Extract cell formulas and values
        for row in worksheet.iter_rows():
            for cell in row:
                cell_address = f"{cell.column_letter}{cell.row}"
                
                # Store cell value
                if cell.value is not None:
                    sheet_data['values'][cell_address] = cell.value
                
                # Extract formula if present
                if cell.data_type == 'f':  # Formula cell
                    formula = cell.value
                    sheet_data['formulas'][cell_address] = formula
                    
                    # Extract formula metadata
                    metadata = self._analyze_formula(formula, cell_address)
                    sheet_data['formula_metadata'][cell_address] = metadata
        
        # Extract named ranges
        if hasattr(worksheet.parent, 'defined_names'):
            for named_range in worksheet.parent.defined_names:
                if named_range.destinations:
                    for title, coordinate in named_range.destinations:
                        if title == worksheet.title:
                            sheet_data['named_ranges'][named_range.name] = coordinate
        
        # Extract data validation rules
        for validation in worksheet.data_validations.dataValidation:
            if validation.sqref:
                sheet_data['data_validation'][str(validation.sqref)] = {
                    'type': validation.type,
                    'formula1': validation.formula1,
                    'formula2': validation.formula2
                }
        
        return sheet_data
    
    def _analyze_formula(self, formula: str, cell_address: str) -> Dict[str, Any]:
        """Analyze a formula and extract metadata"""
        
        metadata = {
            'original_formula': formula,
            'cell_address': cell_address,
            'referenced_cells': [],
            'referenced_ranges': [],
            'excel_functions': [],
            'operators': [],
            'constants': [],
            'formula_type': 'simple',
            'complexity_score': 0
        }
        
        if not formula or not formula.startswith('='):
            return metadata
        
        # Remove the leading '='
        formula_body = formula[1:]
        
        # Extract cell references (e.g., A1, B2:D5)
        cell_pattern = r'([A-Z]+\d+(?::[A-Z]+\d+)?)'
        cell_matches = re.findall(cell_pattern, formula_body)
        
        for match in cell_matches:
            if ':' in match:
                metadata['referenced_ranges'].append(match)
            else:
                metadata['referenced_cells'].append(match)
        
        # Extract Excel functions
        function_pattern = r'([A-Z]+)\s*\('
        function_matches = re.findall(function_pattern, formula_body)
        metadata['excel_functions'] = list(set(function_matches))
        
        # Extract operators
        operator_pattern = r'([\+\-\*\/\^\<\>\=])'
        operator_matches = re.findall(operator_pattern, formula_body)
        metadata['operators'] = list(set(operator_matches))
        
        # Extract numeric constants
        constant_pattern = r'\b(\d+\.?\d*)\b'
        constant_matches = re.findall(constant_pattern, formula_body)
        metadata['constants'] = [float(c) for c in constant_matches]
        
        # Determine formula type
        if len(metadata['excel_functions']) > 0:
            metadata['formula_type'] = 'function'
        elif len(metadata['referenced_cells']) > 5 or len(metadata['referenced_ranges']) > 0:
            metadata['formula_type'] = 'complex_reference'
        elif len(metadata['operators']) > 3:
            metadata['formula_type'] = 'complex_arithmetic'
        else:
            metadata['formula_type'] = 'simple'
        
        # Calculate complexity score
        complexity = (len(metadata['referenced_cells']) + 
                     len(metadata['referenced_ranges']) * 2 + 
                     len(metadata['excel_functions']) * 3 + 
                     len(metadata['operators']))
        metadata['complexity_score'] = complexity
        
        return metadata
    
    def _analyze_cross_sheet_dependencies(self, sheets_data: Dict[str, Any]) -> Dict[str, List[str]]:
        """Analyze dependencies between sheets"""
        
        dependencies = {}
        
        for sheet_name, sheet_data in sheets_data.items():
            sheet_dependencies = []
            
            for cell_address, formula in sheet_data.get('formulas', {}).items():
                # Look for cross-sheet references (e.g., Sheet1!A1)
                cross_sheet_pattern = r"(['\w\s]+)!([A-Z]+\d+(?::[A-Z]+\d+)?)"
                matches = re.findall(cross_sheet_pattern, formula)
                
                for referenced_sheet, referenced_cell in matches:
                    referenced_sheet = referenced_sheet.strip("'")
                    if referenced_sheet != sheet_name and referenced_sheet not in sheet_dependencies:
                        sheet_dependencies.append(referenced_sheet)
            
            if sheet_dependencies:
                dependencies[sheet_name] = sheet_dependencies
        
        return dependencies
    
    def evaluate_formula(self, formula: str, context: Dict[str, Any] = None) -> Union[float, str, bool, None]:
        """Evaluate a formula with given context"""
        
        if not formula or not formula.startswith('='):
            return formula
        
        try:
            # Remove the leading '='
            formula_body = formula[1:]
            
            # Replace cell references with values from context
            if context:
                formula_body = self._substitute_cell_references(formula_body, context)
            
            # Replace Excel functions with Python equivalents
            formula_body = self._replace_excel_functions(formula_body)
            
            # Evaluate the formula
            result = self._safe_eval(formula_body)
            return result
            
        except Exception as e:
            return f"#ERROR: {str(e)}"
    
    def _substitute_cell_references(self, formula: str, context: Dict[str, Any]) -> str:
        """Substitute cell references with actual values"""
        
        # Handle single cell references (e.g., A1, B2)
        def replace_cell(match):
            cell_ref = match.group(0)
            if cell_ref in context:
                value = context[cell_ref]
                if isinstance(value, str):
                    return f'"{value}"'
                return str(value)
            return "0"  # Default value for missing references
        
        cell_pattern = r'\b[A-Z]+\d+\b'
        formula = re.sub(cell_pattern, replace_cell, formula)
        
        return formula
    
    def _replace_excel_functions(self, formula: str) -> str:
        """Replace Excel functions with Python equivalents"""
        
        # Simple replacements for common functions
        replacements = {
            'SUM(': 'sum([',
            'AVERAGE(': 'np.mean([',
            'MAX(': 'max([',
            'MIN(': 'min([',
            'SQRT(': 'np.sqrt(',
            'ABS(': 'abs(',
            'POWER(': 'pow(',
            'EXP(': 'np.exp(',
            'LN(': 'np.log(',
            'LOG(': 'np.log10(',
            'PI()': 'np.pi',
            'TRUE': 'True',
            'FALSE': 'False'
        }
        
        for excel_func, python_func in replacements.items():
            formula = formula.replace(excel_func, python_func)
            
            # Close array brackets for aggregate functions
            if python_func.endswith('['):
                formula = formula.replace(')', '])')
        
        return formula
    
    def _safe_eval(self, expression: str) -> Union[float, str, bool, None]:
        """Safely evaluate a mathematical expression"""
        
        try:
            # Create a safe namespace
            safe_dict = {
                "__builtins__": {},
                "abs": abs,
                "max": max,
                "min": min,
                "sum": sum,
                "pow": pow,
                "round": round,
                "np": np,
                "True": True,
                "False": False
            }
            
            # Evaluate the expression
            result = eval(expression, safe_dict)
            return result
            
        except Exception as e:
            raise ValueError(f"Cannot evaluate expression '{expression}': {str(e)}")
    
    def _excel_if(self, condition, true_value, false_value):
        """Excel IF function implementation"""
        return true_value if condition else false_value
    
    def _excel_and(self, *conditions):
        """Excel AND function implementation"""
        return all(conditions)
    
    def _excel_or(self, *conditions):
        """Excel OR function implementation"""
        return any(conditions)
    
    def extract_engineering_formulas(self, formula_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract and categorize engineering formulas"""
        
        engineering_categories = {
            'structural': {
                'keywords': ['moment', 'shear', 'stress', 'strain', 'deflection', 'force', 'load'],
                'formulas': []
            },
            'hydraulic': {
                'keywords': ['discharge', 'velocity', 'flow', 'pressure', 'head', 'regime', 'afflux'],
                'formulas': []
            },
            'geotechnical': {
                'keywords': ['bearing', 'settlement', 'earth', 'pressure', 'friction', 'cohesion'],
                'formulas': []
            },
            'material': {
                'keywords': ['concrete', 'steel', 'strength', 'modulus', 'grade', 'density'],
                'formulas': []
            },
            'geometric': {
                'keywords': ['area', 'volume', 'length', 'width', 'height', 'thickness', 'span'],
                'formulas': []
            }
        }
        
        # Categorize formulas based on keywords
        for sheet_name, sheet_data in formula_data.get('sheets', {}).items():
            for cell_address, formula in sheet_data.get('formulas', {}).items():
                metadata = sheet_data.get('formula_metadata', {}).get(cell_address, {})
                
                # Check which category this formula belongs to
                for category, category_data in engineering_categories.items():
                    keywords = category_data['keywords']
                    
                    # Check if any keyword appears in the formula or nearby cells
                    formula_lower = formula.lower()
                    if any(keyword in formula_lower for keyword in keywords):
                        category_data['formulas'].append({
                            'sheet': sheet_name,
                            'cell': cell_address,
                            'formula': formula,
                            'metadata': metadata
                        })
        
        return engineering_categories
    
    def validate_formula_integrity(self, formula_data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate the integrity of extracted formulas"""
        
        validation_results = {
            'syntax_errors': [],
            'circular_references': [],
            'missing_references': [],
            'inconsistent_units': [],
            'validation_summary': {
                'total_formulas': 0,
                'valid_formulas': 0,
                'error_formulas': 0,
                'warning_formulas': 0
            }
        }
        
        # Track all cell addresses across sheets
        all_cells = set()
        for sheet_data in formula_data.get('sheets', {}).values():
            all_cells.update(sheet_data.get('values', {}).keys())
            all_cells.update(sheet_data.get('formulas', {}).keys())
        
        # Validate each formula
        for sheet_name, sheet_data in formula_data.get('sheets', {}).items():
            for cell_address, formula in sheet_data.get('formulas', {}).items():
                validation_results['validation_summary']['total_formulas'] += 1
                
                # Check syntax
                if not self._validate_formula_syntax(formula):
                    validation_results['syntax_errors'].append({
                        'sheet': sheet_name,
                        'cell': cell_address,
                        'formula': formula,
                        'error': 'Invalid syntax'
                    })
                    validation_results['validation_summary']['error_formulas'] += 1
                    continue
                
                # Check for missing references
                metadata = sheet_data.get('formula_metadata', {}).get(cell_address, {})
                missing_refs = []
                
                for ref_cell in metadata.get('referenced_cells', []):
                    if ref_cell not in all_cells:
                        missing_refs.append(ref_cell)
                
                if missing_refs:
                    validation_results['missing_references'].append({
                        'sheet': sheet_name,
                        'cell': cell_address,
                        'formula': formula,
                        'missing_references': missing_refs
                    })
                    validation_results['validation_summary']['warning_formulas'] += 1
                else:
                    validation_results['validation_summary']['valid_formulas'] += 1
        
        # Check for circular references
        circular_refs = self._detect_circular_references(formula_data)
        validation_results['circular_references'] = circular_refs
        
        return validation_results
    
    def _validate_formula_syntax(self, formula: str) -> bool:
        """Validate basic formula syntax"""
        
        if not formula or not formula.startswith('='):
            return False
        
        # Check for balanced parentheses
        open_count = formula.count('(')
        close_count = formula.count(')')
        if open_count != close_count:
            return False
        
        # Check for valid characters
        valid_chars = set('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789+-*/^().,!:;<>=&% ')
        if not all(c in valid_chars for c in formula):
            return False
        
        return True
    
    def _detect_circular_references(self, formula_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Detect circular references in formulas"""
        
        circular_refs = []
        
        # Build dependency graph
        dependencies = {}
        for sheet_name, sheet_data in formula_data.get('sheets', {}).items():
            for cell_address, formula in sheet_data.get('formulas', {}).items():
                metadata = sheet_data.get('formula_metadata', {}).get(cell_address, {})
                full_address = f"{sheet_name}!{cell_address}"
                dependencies[full_address] = []
                
                # Add referenced cells
                for ref_cell in metadata.get('referenced_cells', []):
                    dependencies[full_address].append(f"{sheet_name}!{ref_cell}")
        
        # Detect cycles using DFS
        def has_cycle(node, visited, rec_stack, path):
            visited.add(node)
            rec_stack.add(node)
            path.append(node)
            
            for neighbor in dependencies.get(node, []):
                if neighbor not in visited:
                    if has_cycle(neighbor, visited, rec_stack, path):
                        return True
                elif neighbor in rec_stack:
                    # Found a cycle
                    cycle_start = path.index(neighbor)
                    cycle = path[cycle_start:] + [neighbor]
                    circular_refs.append({
                        'cycle': cycle,
                        'description': f"Circular reference detected: {' -> '.join(cycle)}"
                    })
                    return True
            
            path.pop()
            rec_stack.remove(node)
            return False
        
        visited = set()
        for node in dependencies:
            if node not in visited:
                has_cycle(node, visited, set(), [])
        
        return circular_refs
    
    def create_formula_documentation(self, formula_data: Dict[str, Any]) -> str:
        """Create documentation for extracted formulas"""
        
        doc = "# Bridge Design Formula Documentation\n\n"
        doc += f"Generated on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        
        # Summary
        total_formulas = formula_data.get('formula_summary', {}).get('total_formulas', 0)
        doc += f"## Summary\n"
        doc += f"- Total formulas extracted: {total_formulas}\n"
        doc += f"- Number of sheets: {len(formula_data.get('sheets', {}))}\n\n"
        
        # Engineering categories
        engineering_formulas = self.extract_engineering_formulas(formula_data)
        doc += "## Engineering Formula Categories\n\n"
        
        for category, category_data in engineering_formulas.items():
            formulas = category_data['formulas']
            if formulas:
                doc += f"### {category.title()} ({len(formulas)} formulas)\n\n"
                
                for formula_info in formulas[:5]:  # Show first 5 formulas
                    doc += f"**{formula_info['sheet']}!{formula_info['cell']}**\n"
                    doc += f"```\n{formula_info['formula']}\n```\n\n"
                
                if len(formulas) > 5:
                    doc += f"... and {len(formulas) - 5} more formulas\n\n"
        
        # Dependencies
        dependencies = formula_data.get('formula_summary', {}).get('dependencies', {})
        if dependencies:
            doc += "## Cross-Sheet Dependencies\n\n"
            for sheet, deps in dependencies.items():
                doc += f"- **{sheet}** depends on: {', '.join(deps)}\n"
        
        return doc
