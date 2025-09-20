#!/usr/bin/env python3
"""
Generate Sample Excel Files for Bridge Design Application
Creates template Excel files with realistic bridge design data and formulas
Based on Bundan River Bridge TAD project structure
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import math

def create_stability_analysis_excel():
    """Create stability analysis Excel file"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stability Analysis"
    
    # Headers
    headers = {
        'A1': 'Bundan River Bridge - Stability Analysis',
        'A3': 'Project Parameters',
        'A5': 'Parameter',
        'B5': 'Value',
        'C5': 'Unit',
        'A6': 'Structure Height',
        'B6': 6.5,
        'C6': 'm',
        'A7': 'Structure Width',
        'B7': 7.0,
        'C7': 'm',
        'A8': 'Concrete Density',
        'B8': 24.0,
        'C8': 'kN/m³',
        'A9': 'Soil Unit Weight',
        'B9': 18.0,
        'C9': 'kN/m³',
        'A10': 'Angle of Friction',
        'B10': 30.0,
        'C10': 'degrees',
        'A11': 'Bearing Capacity',
        'B11': 450.0,
        'C11': 'kN/m²',
        'A13': 'Load Calculations',
        'A15': 'Self Weight',
        'B15': '=B6*B7*B8',
        'C15': 'kN/m',
        'A16': 'Earth Pressure Coefficient Ka',
        'B16': '=TAN(RADIANS(45-B10/2))^2',
        'C16': '-',
        'A17': 'Active Earth Pressure',
        'B17': '=0.5*B16*B9*B6*B6',
        'C17': 'kN/m',
        'A18': 'Overturning Moment',
        'B18': '=B17*B6/3',
        'C18': 'kN-m/m',
        'A19': 'Resisting Moment',
        'B19': '=B15*B7/2',
        'C19': 'kN-m/m',
        'A20': 'Overturning Factor',
        'B20': '=B19/B18',
        'C20': '-',
        'A21': 'Sliding Factor',
        'B21': '=(B15*0.5)/B17',
        'C21': '-',
        'A22': 'Max Soil Pressure',
        'B22': '=B15/B7*(1+6*B18/(B15*B7))',
        'C22': 'kN/m²',
        'A24': 'Safety Check',
        'A25': 'Overturning Safe?',
        'B25': '=IF(B20>=2,"SAFE","UNSAFE")',
        'A26': 'Sliding Safe?',
        'B26': '=IF(B21>=1.5,"SAFE","UNSAFE")',
        'A27': 'Bearing Safe?',
        'B27': '=IF(B22<=B11,"SAFE","UNSAFE")'
    }
    
    # Apply data to worksheet
    for cell, value in headers.items():
        ws[cell] = value
    
    # Add more detailed force calculations in a separate sheet
    ws2 = wb.create_sheet("Force Details")
    force_data = {
        'A1': 'Detailed Force Calculations',
        'A3': 'Forces Acting on Structure',
        'A5': 'Force Type',
        'B5': 'Magnitude',
        'C5': 'Arm',
        'D5': 'Moment',
        'E5': 'Unit',
        'A6': 'Self Weight',
        'B6': f'={ws.title}!B15',
        'C6': f'={ws.title}!B7/2',
        'D6': '=B6*C6',
        'E6': 'kN-m/m',
        'A7': 'Earth Pressure',
        'B7': f'={ws.title}!B17',
        'C7': f'={ws.title}!B6/3',
        'D7': '=B7*C7',
        'E7': 'kN-m/m',
        'A9': 'Summary',
        'A10': 'Total Vertical Force',
        'B10': '=B6',
        'A11': 'Total Horizontal Force',
        'B11': '=B7',
        'A12': 'Net Overturning Moment',
        'B12': '=D7',
        'A13': 'Net Resisting Moment',
        'B13': '=D6'
    }
    
    for cell, value in force_data.items():
        ws2[cell] = value
    
    return wb

def create_hydraulic_analysis_excel():
    """Create hydraulic analysis Excel file"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Hydraulic Analysis"
    
    hydraulic_data = {
        'A1': 'Bundan River Bridge - Hydraulic Analysis',
        'A3': 'Design Parameters',
        'A5': 'Parameter',
        'B5': 'Value',
        'C5': 'Unit',
        'A6': 'Design Discharge',
        'B6': 902.15,
        'C6': 'cumecs',
        'A7': 'High Flood Level',
        'B7': 101.2,
        'C7': 'm',
        'A8': 'Bed Slope',
        'B8': '1 in 975',
        'C8': '-',
        'A9': 'Manning n',
        'B9': 0.033,
        'C9': '-',
        'A10': 'Silt Factor',
        'B10': 1.5,
        'C10': '-',
        'A11': 'Design Velocity',
        'B11': 3.5,
        'C11': 'm/s',
        'A12': 'Bridge Opening',
        'B12': 75.0,
        'C12': 'm',
        'A13': 'Allowable Afflux',
        'B13': 0.3,
        'C13': 'm',
        'A15': 'Lacey Regime Calculations',
        'A16': 'Regime Width',
        'B16': '=4.75*SQRT(B6)',
        'C16': 'm',
        'A17': 'Regime Depth',
        'B17': '=0.473*(B6/B10)^(1/3)',
        'C17': 'm',
        'A18': 'Regime Velocity',
        'B18': '=1.17*SQRT(B10)',
        'C18': 'm/s',
        'A19': 'Regime Area',
        'B19': '=B16*B17',
        'C19': 'm²',
        'A21': 'Afflux Calculations',
        'A22': 'Approach Velocity',
        'B22': '=B6/B19',
        'C22': 'm/s',
        'A23': 'Bridge Velocity',
        'B23': '=B6/(B12*B17)',
        'C23': 'm/s',
        'A24': 'Molesworth Afflux',
        'B24': '=1.8*B6^2/(2*9.81*0.95^2*B12^2)',
        'C24': 'm',
        'A25': 'Energy Afflux',
        'B25': '=(B23^2-B22^2)/(2*9.81)',
        'C25': 'm',
        'A26': 'Design Afflux',
        'B26': '=MAX(B24,B25)',
        'C26': 'm',
        'A28': 'Scour Analysis',
        'A29': 'Discharge per unit width',
        'B29': '=B6/B16',
        'C29': 'cumecs/m',
        'A30': 'Lacey Scour Depth',
        'B30': '=0.473*(B29^2/B10)^(1/3)',
        'C30': 'm',
        'A31': 'Design Scour with Safety',
        'B31': '=B30*1.5',
        'C31': 'm',
        'A32': 'Foundation Level Required',
        'B32': '=B7-B31',
        'C32': 'm',
        'A34': 'Adequacy Check',
        'A35': 'Waterway Ratio',
        'B35': '=B12/B16',
        'C35': '-',
        'A36': 'Afflux Check',
        'B36': '=IF(B26<=B13,"ADEQUATE","INADEQUATE")',
        'A37': 'Velocity Check',
        'B37': '=IF(B23<=B11,"SAFE","HIGH")'
    }
    
    for cell, value in hydraulic_data.items():
        ws[cell] = value
    
    # Add scour details sheet
    ws2 = wb.create_sheet("Scour Details")
    scour_data = {
        'A1': 'Detailed Scour Analysis',
        'A3': 'Multiple Methods for Scour Depth',
        'A5': 'Method',
        'B5': 'Formula',
        'C5': 'Result',
        'D5': 'Unit',
        'A6': 'Lacey Method',
        'B6': '0.473*(q²/f)^(1/3)',
        'C6': f'={ws.title}!B30',
        'D6': 'm',
        'A7': 'Blench Method',
        'B7': '1.35*(Q/f)^(1/3)',
        'C7': f'=1.35*({ws.title}!B6/{ws.title}!B10)^(1/3)',
        'D7': 'm',
        'A8': 'IRC Method',
        'B8': '2.0*R*(V/Vc)²',
        'C8': f'=2*{ws.title}!B17*({ws.title}!B18/{ws.title}!B18)^2',
        'D8': 'm',
        'A10': 'Design Scour',
        'B10': 'Maximum of above',
        'C10': '=MAX(C6:C8)',
        'D10': 'm',
        'A11': 'With Safety Factor',
        'B11': 'Design × 1.5',
        'C11': '=C10*1.5',
        'D11': 'm'
    }
    
    for cell, value in scour_data.items():
        ws2[cell] = value
    
    return wb

def create_cross_section_excel():
    """Create cross section design Excel file"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cross Section Design"
    
    cross_section_data = {
        'A1': 'Bridge Cross Section Design',
        'A3': 'Geometric Properties',
        'A5': 'Parameter',
        'B5': 'Value',
        'C5': 'Unit',
        'A6': 'Total Width',
        'B6': 9.5,
        'C6': 'm',
        'A7': 'Carriageway Width',
        'B7': 7.5,
        'C7': 'm',
        'A8': 'Footpath Width (each)',
        'B8': 1.0,
        'C8': 'm',
        'A9': 'Slab Thickness',
        'B9': 0.5,
        'C9': 'm',
        'A10': 'Wearing Coat',
        'B10': 0.075,
        'C10': 'm',
        'A11': 'Effective Depth',
        'B11': '=B9-0.05',
        'C11': 'm',
        'A13': 'Material Properties',
        'A14': 'Concrete Grade',
        'B14': 'M25',
        'A15': 'fck',
        'B15': 25,
        'C15': 'N/mm²',
        'A16': 'Steel Grade',
        'B16': 'Fe415',
        'A17': 'fy',
        'B17': 415,
        'C17': 'N/mm²',
        'A18': 'Concrete Density',
        'B18': 25,
        'C18': 'kN/m³',
        'A20': 'Load Calculations',
        'A21': 'Self Weight',
        'B21': '=B9*B6*B18',
        'C21': 'kN/m',
        'A22': 'Wearing Coat Load',
        'B22': '=B10*B7*22',
        'C22': 'kN/m',
        'A23': 'Crash Barrier Load',
        'B23': 6.0,
        'C23': 'kN/m',
        'A24': 'Total Dead Load',
        'B24': '=B21+B22+B23',
        'C24': 'kN/m',
        'A25': 'Live Load (Class A)',
        'B25': '=5*B7',
        'C25': 'kN/m',
        'A26': 'Impact Factor',
        'B26': 0.25,
        'C26': '-',
        'A27': 'Live Load with Impact',
        'B27': '=B25*(1+B26)',
        'C27': 'kN/m',
        'A29': 'Design Forces (10m span)',
        'A30': 'Factored Dead Load',
        'B30': '=B24*1.35',
        'C30': 'kN/m',
        'A31': 'Factored Live Load',
        'B31': '=B27*1.75',
        'C31': 'kN/m',
        'A32': 'Total Design Load',
        'B32': '=B30+B31',
        'C32': 'kN/m',
        'A33': 'Design Moment',
        'B33': '=B32*10^2/8',
        'C33': 'kN-m/m',
        'A34': 'Design Shear',
        'B34': '=B32*10/2',
        'C34': 'kN/m',
        'A36': 'Reinforcement Design',
        'A37': 'Required Steel Area',
        'B37': '=B33*1000000/(0.87*B17*0.9*B11*1000)',
        'C37': 'mm²/m',
        'A38': 'Minimum Steel (0.12%)',
        'B38': '=0.12*1000*B9*1000/100',
        'C38': 'mm²/m',
        'A39': 'Provided Steel Area',
        'B39': '=MAX(B37,B38)',
        'C39': 'mm²/m',
        'A40': '16mm bars spacing',
        'B40': '=201*1000/B39',
        'C40': 'mm',
        'A41': 'Provided Spacing',
        'B41': '=ROUND(B40/25,0)*25',
        'C41': 'mm',
        'A42': 'Actual Steel Provided',
        'B42': '=201*1000/B41',
        'C42': 'mm²/m',
        'A44': 'Design Check',
        'A45': 'Steel Ratio',
        'B45': '=B42/(1000*B11*1000)*100',
        'C45': '%',
        'A46': 'Deflection Check',
        'B46': '=IF(B45<1,"PASS","CHECK")',
        'A47': 'Shear Check',
        'B47': '=IF(B34*1000/(1000*B11*1000)<=0.62*SQRT(B15),"SAFE","STIRRUPS REQD")'
    }
    
    for cell, value in cross_section_data.items():
        ws[cell] = value
    
    # Add reinforcement details sheet
    ws2 = wb.create_sheet("Reinforcement Details")
    rebar_data = {
        'A1': 'Reinforcement Layout Details',
        'A3': 'Main Reinforcement',
        'A5': 'Bar Size',
        'B5': 'Spacing',
        'C5': 'Area per meter',
        'D5': 'Location',
        'A6': '16mm',
        'B6': f'={ws.title}!B41',
        'C6': f'={ws.title}!B42',
        'D6': 'Bottom',
        'A7': '12mm',
        'B7': 250,
        'C7': '=113*1000/B7',
        'D7': 'Top',
        'A9': 'Distribution Steel',
        'A10': '12mm @ 250mm c/c',
        'A11': 'Both directions',
        'A13': 'Shear Reinforcement',
        'A14': 'Check required',
        'B14': f'={ws.title}!B47',
        'A15': 'If required: 8mm stirrups',
        'A16': 'Spacing as per design'
    }
    
    for cell, value in rebar_data.items():
        ws2[cell] = value
    
    return wb

def create_abutment_design_excel():
    """Create abutment design Excel file"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Abutment Design"
    
    abutment_data = {
        'A1': 'Bridge Abutment Design - Type 1 Battered',
        'A3': 'Geometric Parameters',
        'A5': 'Parameter',
        'B5': 'Value',
        'C5': 'Unit',
        'A6': 'Total Height',
        'B6': 6.5,
        'C6': 'm',
        'A7': 'Stem Top Thickness',
        'B7': 0.5,
        'C7': 'm',
        'A8': 'Stem Base Thickness',
        'B8': 1.0,
        'C8': 'm',
        'A9': 'Base Length',
        'B9': 5.0,
        'C9': 'm',
        'A10': 'Base Width',
        'B10': 8.0,
        'C10': 'm',
        'A11': 'Heel Length',
        'B11': 3.0,
        'C11': 'm',
        'A12': 'Toe Length',
        'B12': 2.0,
        'C12': 'm',
        'A14': 'Material Properties',
        'A15': 'Concrete Density',
        'B15': 25,
        'C15': 'kN/m³',
        'A16': 'Soil Unit Weight',
        'B16': 18,
        'C16': 'kN/m³',
        'A17': 'Angle of Friction',
        'B17': 30,
        'C17': 'degrees',
        'A18': 'Bearing Capacity',
        'B18': 450,
        'C18': 'kN/m²',
        'A19': 'Surcharge Load',
        'B19': 10,
        'C19': 'kN/m²',
        'A21': 'Load Calculations',
        'A22': 'Stem Volume',
        'B22': '=0.5*(B7+B8)*B6*B10',
        'C22': 'm³/m',
        'A23': 'Stem Weight',
        'B23': '=B22*B15',
        'C23': 'kN/m',
        'A24': 'Base Volume',
        'B24': '=B9*B10*B8',
        'C24': 'm³/m',
        'A25': 'Base Weight',
        'B25': '=B24*B15',
        'C25': 'kN/m',
        'A26': 'Total Dead Load',
        'B26': '=B23+B25',
        'C26': 'kN/m',
        'A27': 'Bridge Reaction',
        'B27': 200,
        'C27': 'kN/m',
        'A28': 'Total Vertical Load',
        'B28': '=B26+B27',
        'C28': 'kN/m',
        'A30': 'Earth Pressure',
        'A31': 'Ka (Active)',
        'B31': '=TAN(RADIANS(45-B17/2))^2',
        'C31': '-',
        'A32': 'Active Pressure at Base',
        'B32': '=B31*B16*B6',
        'C32': 'kN/m²',
        'A33': 'Total Active Force',
        'B33': '=0.5*B32*B6',
        'C33': 'kN/m',
        'A34': 'Surcharge Force',
        'B34': '=B31*B19*B6',
        'C34': 'kN/m',
        'A35': 'Total Horizontal Force',
        'B35': '=B33+B34',
        'C35': 'kN/m',
        'A37': 'Stability Checks',
        'A38': 'Overturning Moment',
        'B38': '=B33*B6/3+B34*B6/2',
        'C38': 'kN-m/m',
        'A39': 'Resisting Moment',
        'B39': '=B26*B9/2+B27*B12',
        'C39': 'kN-m/m',
        'A40': 'Overturning Factor',
        'B40': '=B39/B38',
        'C40': '-',
        'A41': 'Sliding Resistance',
        'B41': '=B28*0.5',
        'C41': 'kN/m',
        'A42': 'Sliding Factor',
        'B42': '=B41/B35',
        'C42': '-',
        'A44': 'Bearing Pressure',
        'A45': 'Eccentricity',
        'B45': '=B38/B28',
        'C45': 'm',
        'A46': 'Max Pressure',
        'B46': '=B28/B9*(1+6*B45/B9)',
        'C46': 'kN/m²',
        'A47': 'Min Pressure',
        'B47': '=B28/B9*(1-6*B45/B9)',
        'C47': 'kN/m²',
        'A49': 'Safety Assessment',
        'A50': 'Overturning Safe?',
        'B50': '=IF(B40>=2,"SAFE","UNSAFE")',
        'A51': 'Sliding Safe?',
        'B51': '=IF(B42>=1.5,"SAFE","UNSAFE")',
        'A52': 'Bearing Safe?',
        'B52': '=IF(B46<=B18,"SAFE","UNSAFE")',
        'A53': 'Overall Status',
        'B53': '=IF(AND(B50="SAFE",B51="SAFE",B52="SAFE"),"SAFE","REVIEW REQUIRED")'
    }
    
    for cell, value in abutment_data.items():
        ws[cell] = value
    
    # Add foundation design sheet
    ws2 = wb.create_sheet("Foundation Design")
    foundation_data = {
        'A1': 'Foundation Design Details',
        'A3': 'Foundation Type Assessment',
        'A5': 'Required Foundation Depth',
        'B5': f'={ws.title}!B6+2',
        'C5': 'm',
        'A6': 'Bearing Pressure Check',
        'B6': f'={ws.title}!B52',
        'A7': 'Foundation Type',
        'B7': '=IF(B6="SAFE","Shallow Foundation","Deep Foundation Required")',
        'A9': 'Shallow Foundation Design',
        'A10': 'Foundation Width',
        'B10': f'={ws.title}!B9',
        'C10': 'm',
        'A11': 'Foundation Depth',
        'B11': 2.0,
        'C11': 'm',
        'A12': 'Foundation Volume',
        'B12': f'=B10*{ws.title}!B10*B11',
        'C12': 'm³/m',
        'A13': 'Concrete Required',
        'B13': '=B12',
        'C13': 'm³/m',
        'A15': 'Reinforcement Design',
        'A16': 'Foundation Steel',
        'B16': '=0.15*B13*1000*25',
        'C16': 'kg/m',
        'A17': 'Main Bars',
        'B17': '16mm @ 200mm c/c both ways',
        'A18': 'Distribution Bars',
        'B18': '12mm @ 250mm c/c both ways'
    }
    
    for cell, value in foundation_data.items():
        ws2[cell] = value
    
    return wb

def apply_formatting(wb):
    """Apply basic formatting to Excel workbook"""
    
    for ws in wb.worksheets:
        # Header formatting
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Apply to first row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Format parameter headers
        param_font = Font(bold=True, size=10)
        for row in ws.iter_rows(min_row=5, max_row=5):
            for cell in row:
                if cell.value:
                    cell.font = param_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

def main():
    """Generate all sample Excel files"""
    
    print("Creating sample Excel files for Bridge Design Application...")
    
    # Create templates directory if it doesn't exist
    os.makedirs("templates", exist_ok=True)
    
    # Generate Excel files
    files_to_create = [
        ("Stability_Analysis_Template.xlsx", create_stability_analysis_excel),
        ("Hydraulic_Analysis_Template.xlsx", create_hydraulic_analysis_excel),
        ("Cross_Section_Design_Template.xlsx", create_cross_section_excel),
        ("Abutment_Design_Template.xlsx", create_abutment_design_excel)
    ]
    
    for filename, create_func in files_to_create:
        try:
            print(f"Creating {filename}...")
            wb = create_func()
            apply_formatting(wb)
            
            # Save to templates directory
            filepath = os.path.join("templates", filename)
            wb.save(filepath)
            print(f"✓ Successfully created {filepath}")
            
        except Exception as e:
            print(f"✗ Error creating {filename}: {str(e)}")
    
    print("\nSample Excel files created successfully!")
    print("These files contain realistic bridge design data and formulas")
    print("that match the structure expected by the Bridge Design Application.")

if __name__ == "__main__":
    main()