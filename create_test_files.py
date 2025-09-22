#!/usr/bin/env python3
"""
Simple script to create test Excel files for the Bridge Design Application
"""

import sys
import os

# Add the current directory to the path to import modules
sys.path.insert(0, os.path.dirname(__file__))

try:
    import openpyxl
    from openpyxl import Workbook
    print("✅ Successfully imported required modules")
except ImportError as e:
    print(f"❌ Failed to import modules: {e}")
    sys.exit(1)

def create_test_stability_file(filename):
    """Create a simple stability analysis Excel file"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise Exception("Failed to create worksheet")
    
    ws.title = "Stability Analysis"
    
    # Add headers
    ws['A1'] = 'Bridge Stability Analysis - Test File'
    ws['A3'] = 'Parameters'
    ws['A4'] = 'Structure Height'
    ws['B4'] = 5.0
    ws['C4'] = 'm'
    
    ws['A5'] = 'Structure Width'
    ws['B5'] = 6.0
    ws['C5'] = 'm'
    
    ws['A6'] = 'Concrete Density'
    ws['B6'] = 24.0
    ws['C6'] = 'kN/m³'
    
    # Add calculations
    ws['A8'] = 'Calculations'
    ws['A9'] = 'Self Weight'
    ws['B9'] = '=B4*B5*B6'
    ws['C9'] = 'kN/m'
    
    ws['A10'] = 'Overturning Factor'
    ws['B10'] = '=B9*0.5'  # Simplified calculation
    ws['C10'] = '-'
    
    ws['A11'] = 'Safety Check'
    ws['B11'] = '=IF(B10>=2,"SAFE","CHECK")'
    
    # Save file
    wb.save(filename)
    print(f"✅ Created {filename}")

def create_test_hydraulic_file(filename):
    """Create a simple hydraulic analysis Excel file"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise Exception("Failed to create worksheet")
    
    ws.title = "Hydraulic Analysis"
    
    # Add headers
    ws['A1'] = 'Hydraulic Analysis - Test File'
    ws['A3'] = 'Parameters'
    ws['A4'] = 'Design Discharge'
    ws['B4'] = 800.0
    ws['C4'] = 'cumecs'
    
    ws['A5'] = 'Bridge Opening'
    ws['B5'] = 70.0
    ws['C5'] = 'm'
    
    # Add calculations
    ws['A7'] = 'Calculations'
    ws['A8'] = 'Flow Velocity'
    ws['B8'] = '=B4/B5'
    ws['C8'] = 'm/s'
    
    ws['A9'] = 'Scour Depth'
    ws['B9'] = '=0.5*B8'
    ws['C9'] = 'm'
    
    ws['A10'] = 'Adequacy'
    ws['B10'] = '=IF(B9<2,"ADEQUATE","INADEQUATE")'
    
    # Save file
    wb.save(filename)
    print(f"✅ Created {filename}")

def create_test_abutment_file(filename):
    """Create a simple abutment design Excel file"""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise Exception("Failed to create worksheet")
    
    ws.title = "Abutment Design"
    
    # Add headers
    ws['A1'] = 'Abutment Design - Test File'
    ws['A3'] = 'Parameters'
    ws['A4'] = 'Abutment Height'
    ws['B4'] = 4.5
    ws['C4'] = 'm'
    
    ws['A5'] = 'Backfill Angle'
    ws['B5'] = 30.0
    ws['C5'] = 'degrees'
    
    # Add calculations
    ws['A7'] = 'Calculations'
    ws['A8'] = 'Earth Pressure'
    ws['B8'] = '=0.5*TAN(RADIANS(45-B5/2))^2'
    ws['C8'] = 'kN/m²'
    
    ws['A9'] = 'Design Status'
    ws['B9'] = '=IF(B8<0.5,"ACCEPTABLE","REVIEW")'
    
    # Save file
    wb.save(filename)
    print(f"✅ Created {filename}")

def main():
    """Create test Excel files"""
    print("Creating test Excel files for Bridge Design Application...")
    
    # Create test directory if it doesn't exist
    os.makedirs("test_files", exist_ok=True)
    
    # Create test files
    test_files = [
        ("test_stability_1.xlsx", create_test_stability_file),
        ("test_hydraulic_1.xlsx", create_test_hydraulic_file),
        ("test_abutment_1.xlsx", create_test_abutment_file)
    ]
    
    for filename, create_func in test_files:
        filepath = os.path.join("test_files", filename)
        try:
            create_func(filepath)
        except Exception as e:
            print(f"❌ Error creating {filename}: {str(e)}")
    
    print("\n✅ Test Excel files created successfully in the 'test_files' directory!")

if __name__ == "__main__":
    main()