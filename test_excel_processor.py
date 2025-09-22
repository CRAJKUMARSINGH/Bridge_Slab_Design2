#!/usr/bin/env python3
"""
Test file for Excel Processor module
"""

import sys
import os
import tempfile

# Add the current directory to the path to import modules
sys.path.insert(0, os.path.dirname(__file__))

try:
    from modules.excel_processor import ExcelProcessor
    import openpyxl
    print("✅ Successfully imported required modules")
except ImportError as e:
    print(f"❌ Failed to import modules: {e}")
    sys.exit(1)

def create_test_excel():
    """Create a simple test Excel file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Test Sheet"
        
        # Add some test data and formulas
        ws['A1'] = "Parameter"
        ws['B1'] = "Value"
        ws['C1'] = "Unit"
        
        ws['A2'] = "Length"
        ws['B2'] = 10.5
        ws['C2'] = "m"
        
        ws['A3'] = "Width"
        ws['B3'] = 5.2
        ws['C3'] = "m"
        
        ws['A4'] = "Area"
        ws['B4'] = "=B2*B3"
        ws['C4'] = "m²"
        
        ws['A5'] = "Safety Check"
        ws['B5'] = "=IF(B4>50,\"OK\",\"Check\")"
    
    # Save to a temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    return temp_file.name

def test_excel_processor():
    """Test the ExcelProcessor class"""
    print("Testing ExcelProcessor...")
    
    # Create test Excel file
    test_file_path = create_test_excel()
    
    try:
        # Initialize processor
        processor = ExcelProcessor()
        
        # Process the Excel file
        with open(test_file_path, 'rb') as f:
            processed_data = processor.process_excel_file(f)
        
        # Check that data was processed correctly
        assert 'sheets' in processed_data, "Processed data should contain sheets"
        assert 'Test Sheet' in processed_data['sheets'], "Should contain 'Test Sheet'"
        
        sheet_data = processed_data['sheets']['Test Sheet']
        assert 'formulas' in sheet_data, "Sheet data should contain formulas"
        assert 'values' in sheet_data, "Sheet data should contain values"
        
        # Check specific formulas
        formulas = sheet_data['formulas']
        assert 'B4' in formulas, "Should contain formula in B4"
        assert 'B5' in formulas, "Should contain formula in B5"
        
        print("✅ ExcelProcessor test passed!")
        print(f"   - Found {len(formulas)} formulas")
        print(f"   - Formula in B4: {formulas['B4']}")
        print(f"   - Formula in B5: {formulas['B5']}")
        
        return True
        
    except Exception as e:
        print(f"❌ ExcelProcessor test failed: {str(e)}")
        return False
        
    finally:
        # Clean up temporary file
        os.unlink(test_file_path)

if __name__ == "__main__":
    success = test_excel_processor()
    sys.exit(0 if success else 1)